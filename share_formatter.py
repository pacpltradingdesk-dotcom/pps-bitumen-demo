"""
Share Formatter — Premium Output Engine
========================================
Centralized, branded output generators for all PPS Anantam sharing channels:

  Text channels
    format_quote_whatsapp(...)      → pretty WhatsApp-ready string
    format_quote_telegram(...)      → Markdown v2 compatible string
    format_payment_reminder(...)    → debt reminder (WA/TG)
    format_followup(...)            → touchpoint follow-up

  HTML / Email
    format_quote_email_html(...)    → branded HTML email body + subject

  Excel
    build_styled_excel(sheets)      → returns bytes, openpyxl-styled

  PDF
    build_quote_pdf(...)            → ReportLab premium quote PDF bytes

Design principles:
  - Company branding in every output (PPS Anantam, bank, GST)
  - Indian number formatting (₹1,23,456) consistent everywhere
  - Unicode box/divider characters in text for visual structure
  - Fallbacks so missing deps never crash the caller
"""
from __future__ import annotations

import datetime as _dt
import io as _io
from typing import Any


# ═══════════════════════════════════════════════════════════════════════════
# Company constants — single source of truth
# ═══════════════════════════════════════════════════════════════════════════

COMPANY = {
    "name":      "PPS Anantams Corporation Pvt Ltd",
    "short":     "PPS Anantams",
    "tagline":   "Har decision data-driven · har rupiya margin-driven",
    "city":      "Vadodara, Gujarat",
    "gst":       "24AAHCV1611L2ZD",
    "pan":       "AAHCV1611L",
    "cin":       "U46632GJ2019PTC110676",
    "owner":     "Prince P Shah",
    "email":     "desk@ppsanantams.com",
    "phone":     "+91 7795242424",
    "bank":      "ICICI Bank",
    "bank_ac":   "184105001402",
    "bank_ifsc": "ICIC0001841",
    "hsn":       "27132000",
    "terms":     "Ex-Terminal / Ex-Warehouse, 100% advance, 24hr validity",
    "logo":      "pps_logo_brand.jpg",  # relative to share_formatter.py
}


# Brand palette
BRAND = {
    "navy":        "#0F1F3C",   # deep navy (header)
    "navy_light":  "#1E3A5F",   # navy accent
    "gold":        "#C9A84C",   # gold accent (labels / dividers)
    "gold_dark":   "#A88430",
    "amber_bg":    "#FFFBEB",   # grand-total band
    "amber_fg":    "#92400E",
    "green_soft":  "#ECFDF5",   # thank-you band
    "green_fg":    "#065F46",
    "grey_50":     "#F9FAFB",
    "grey_100":    "#F3F4F6",
    "grey_200":    "#E5E7EB",
    "grey_500":    "#6B7280",
    "grey_700":    "#374151",
    "white":       "#FFFFFF",
}


# Amount to words (Indian system)
def _num_to_words_indian(n: int) -> str:
    """Convert an integer to Indian English words (lakh/crore). Best-effort."""
    try:
        n = int(n)
    except (TypeError, ValueError):
        return ""
    if n == 0:
        return "Zero"
    below20 = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven",
               "Eight", "Nine", "Ten", "Eleven", "Twelve", "Thirteen",
               "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty",
            "Seventy", "Eighty", "Ninety"]

    def two(x):
        if x < 20:
            return below20[x]
        return tens[x // 10] + (" " + below20[x % 10] if x % 10 else "")

    def three(x):
        if x >= 100:
            return below20[x // 100] + " Hundred" + (" " + two(x % 100) if x % 100 else "")
        return two(x)

    parts = []
    crore = n // 10000000
    if crore:
        parts.append(three(crore) + " Crore")
        n %= 10000000
    lakh = n // 100000
    if lakh:
        parts.append(three(lakh) + " Lakh")
        n %= 100000
    thousand = n // 1000
    if thousand:
        parts.append(three(thousand) + " Thousand")
        n %= 1000
    if n:
        parts.append(three(n))
    return " ".join(parts).strip()


# ═══════════════════════════════════════════════════════════════════════════
# Formatting helpers
# ═══════════════════════════════════════════════════════════════════════════

def fmt_inr(amount) -> str:
    """₹1,23,456 with Indian comma grouping."""
    try:
        n = int(float(amount))
    except (TypeError, ValueError):
        return str(amount)
    s = str(abs(n))
    if len(s) <= 3:
        out = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        groups = []
        while rest:
            groups.insert(0, rest[-2:])
            rest = rest[:-2]
        out = ",".join(groups) + "," + last3
    sign = "-" if n < 0 else ""
    return f"{sign}\u20b9{out}"


def _amount_in_words(amount) -> str:
    """Convert amount to Indian-style words. e.g. 125000 → 'Rupees One Lakh Twenty Five Thousand'."""
    try:
        n = int(round(float(amount)))
    except (TypeError, ValueError):
        return "Rupees Zero"
    if n == 0:
        return "Rupees Zero"
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
            "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def _under_hundred(x):
        if x < 20:
            return ones[x]
        return tens[x // 10] + (" " + ones[x % 10] if x % 10 else "")

    def _under_thousand(x):
        if x < 100:
            return _under_hundred(x)
        return ones[x // 100] + " Hundred" + (" " + _under_hundred(x % 100) if x % 100 else "")

    parts = []
    crore = n // 10000000
    n %= 10000000
    lakh = n // 100000
    n %= 100000
    thousand = n // 1000
    n %= 1000
    rest = n

    if crore:
        parts.append(_under_thousand(crore) + " Crore")
    if lakh:
        parts.append(_under_thousand(lakh) + " Lakh")
    if thousand:
        parts.append(_under_thousand(thousand) + " Thousand")
    if rest:
        parts.append(_under_thousand(rest))
    return "Rupees " + " ".join(parts)


def _today() -> str:
    return _dt.datetime.now().strftime("%d %b %Y")


def _today_long() -> str:
    return _dt.datetime.now().strftime("%d %B %Y")


def _now_time() -> str:
    return _dt.datetime.now().strftime("%H:%M IST")


def _divider(char: str = "─", n: int = 32) -> str:
    return char * n


# ═══════════════════════════════════════════════════════════════════════════
# WhatsApp — premium format
# ═══════════════════════════════════════════════════════════════════════════

def format_quote_whatsapp(customer_name: str, city: str, grade: str,
                          qty_mt: float, price_per_mt: float,
                          source: str = "", savings_pct: float = 0,
                          validity_hours: int = 24,
                          dispatch_hours: int = 48) -> str:
    """
    Premium WhatsApp quote — box-drawn header, clean sections, CTA footer.
    Bold uses WhatsApp *...* syntax.
    """
    total = qty_mt * price_per_mt
    saving_line = ""
    if savings_pct and savings_pct > 0:
        saving_line = f"\n💚 *Market se {savings_pct:.1f}% cheaper*\n"

    src_line = f"\n🏭 Source: {source}" if source else ""

    return (
        f"╔══════════════════╗\n"
        f"║  🏛️ *{COMPANY['short'].upper()}*  ║\n"
        f"║  BITUMEN QUOTATION  ║\n"
        f"╚══════════════════╝\n"
        f"📅 {_today_long()} · {_now_time()}\n"
        f"\n"
        f"*To:* {customer_name}\n"
        f"*Loc:* 📍 {city}"
        f"{src_line}\n"
        f"{_divider('━')}\n"
        f"🏗️  Grade       : *{grade}*\n"
        f"📦  Quantity    : *{qty_mt:,.0f} MT*\n"
        f"💰  Rate        : *{fmt_inr(price_per_mt)}/MT*\n"
        f"🧮  Total Value : *{fmt_inr(total)}*\n"
        f"{saving_line}"
        f"{_divider('━')}\n"
        f"🚚  Dispatch    : Within {dispatch_hours} hrs\n"
        f"⏳  Validity    : {validity_hours} hours only\n"
        f"💳  Payment     : 100% advance\n"
        f"📑  GST         : 18% (HSN {COMPANY['hsn']})\n"
        f"{_divider('━')}\n"
        f"🏦 *Bank Details*\n"
        f"  {COMPANY['bank']}\n"
        f"  A/C: {COMPANY['bank_ac']}\n"
        f"  IFSC: {COMPANY['bank_ifsc']}\n"
        f"{_divider('━')}\n"
        f"✅ *Reply CONFIRM* to lock this price\n"
        f"📞 Call: {COMPANY['phone']}\n"
        f"\n"
        f"— {COMPANY['short']}\n"
        f"   {COMPANY['city']}"
    )


def format_followup_whatsapp(customer_name: str, ref: str = "",
                              days_since: int = 1) -> str:
    """Warm follow-up touchpoint — short & respectful."""
    ref_part = f" (Ref: {ref})" if ref else ""
    return (
        f"🙏 *Namaste {customer_name} ji*,\n"
        f"\n"
        f"Aapko hamari recent bitumen offer{ref_part} pasand aayi?\n"
        f"\n"
        f"Rate abhi bhi valid hai — confirmation par 48 hrs mein\n"
        f"dispatch ho jayega.\n"
        f"\n"
        f"Koi sawaal ho to ek call do 📞 {COMPANY['phone']}\n"
        f"\n"
        f"— {COMPANY['short']}"
    )


def format_reactivation_whatsapp(customer_name: str, city: str,
                                  new_price: float, old_price: float) -> str:
    """Price-drop reactivation message for cold customer."""
    savings = max(0, old_price - new_price)
    return (
        f"╔══════════════════╗\n"
        f"║  🎯 *PRICE DROP ALERT*  ║\n"
        f"╚══════════════════╝\n"
        f"\n"
        f"Dear *{customer_name}*,\n"
        f"\n"
        f"Market rates down hain 📉 — aap ke liye specially:\n"
        f"\n"
        f"🏭 VG-30 at *{fmt_inr(new_price)}/MT* landed {city}\n"
        f"📊 Aapka last rate: {fmt_inr(old_price)}/MT\n"
        f"💰 *Saving: {fmt_inr(savings)}/MT*\n"
        f"\n"
        f"Limited stock — Reply *YES* for formal quote.\n"
        f"\n"
        f"— {COMPANY['short']} · {COMPANY['phone']}"
    )


def format_payment_reminder_whatsapp(customer_name: str, amount: float,
                                      invoice_ref: str = "",
                                      days_overdue: int = 0) -> str:
    """Payment reminder — tone varies by overdue days."""
    if days_overdue <= 0:
        tone = "📌 *Gentle Reminder*"
    elif days_overdue < 7:
        tone = "⏰ *Payment Reminder*"
    elif days_overdue < 15:
        tone = "⚠️ *Urgent — Payment Overdue*"
    else:
        tone = "🚨 *Critical — Immediate Action Required*"

    overdue_line = f"\n📆 Overdue: *{days_overdue} days*" if days_overdue > 0 else ""
    ref_line = f"\n📄 Reference: {invoice_ref}" if invoice_ref else ""

    return (
        f"{tone}\n"
        f"{_divider('━')}\n"
        f"\n"
        f"Dear *{customer_name}*,\n"
        f"\n"
        f"Aap se request hai:\n"
        f"💰 Pending amount: *{fmt_inr(amount)}*{overdue_line}{ref_line}\n"
        f"\n"
        f"Please arrange payment at the earliest.\n"
        f"\n"
        f"🏦 *Bank:* {COMPANY['bank']}\n"
        f"   A/C: `{COMPANY['bank_ac']}`\n"
        f"   IFSC: `{COMPANY['bank_ifsc']}`\n"
        f"   GST: {COMPANY['gst']}\n"
        f"\n"
        f"Payment receipt {COMPANY['phone']} par bhej dijiye.\n"
        f"\n"
        f"Thank you 🙏\n"
        f"— {COMPANY['short']}"
    )


# ═══════════════════════════════════════════════════════════════════════════
# Telegram — MarkdownV2 compatible
# ═══════════════════════════════════════════════════════════════════════════

def _tg_escape(text: str) -> str:
    """Escape MarkdownV2 reserved chars."""
    for ch in r"_*[]()~`>#+-=|{}.!":
        text = text.replace(ch, "\\" + ch)
    return text


def format_quote_telegram(customer_name: str, city: str, grade: str,
                          qty_mt: float, price_per_mt: float,
                          source: str = "") -> str:
    """Telegram MarkdownV2 — same content, escaped properly."""
    total = qty_mt * price_per_mt
    body = (
        f"🏛️ *{COMPANY['short']}* — Bitumen Quotation\n"
        f"📅 {_today_long()}\n\n"
        f"👤 Customer: *{customer_name}*\n"
        f"📍 Location: {city}\n"
        + (f"🏭 Source: {source}\n" if source else "")
        + f"\n"
        f"🏗️ Grade: *{grade}*\n"
        f"📦 Qty: *{qty_mt:,.0f} MT*\n"
        f"💰 Rate: *{fmt_inr(price_per_mt)}/MT*\n"
        f"🧮 Total: *{fmt_inr(total)}*\n\n"
        f"🚚 Dispatch: 48 hrs · ⏳ Validity: 24 hrs\n"
        f"💳 100% advance · GST 18%\n\n"
        f"🏦 {COMPANY['bank']} · A/C {COMPANY['bank_ac']} · IFSC {COMPANY['bank_ifsc']}\n\n"
        f"✅ Reply CONFIRM to lock price\n"
        f"📞 {COMPANY['phone']}"
    )
    return _tg_escape(body)


# ═══════════════════════════════════════════════════════════════════════════
# Email — HTML body
# ═══════════════════════════════════════════════════════════════════════════

def format_quote_email_html(customer_name: str, city: str, grade: str,
                            qty_mt: float, price_per_mt: float,
                            source: str = "") -> dict:
    """Returns {'subject': str, 'body_html': str, 'body_text': str}."""
    total = qty_mt * price_per_mt
    subject = f"Bitumen {grade} — {fmt_inr(price_per_mt)}/MT Landed {city} | {COMPANY['short']}"

    body_html = f"""
<!DOCTYPE html>
<html><body style="font-family:'Inter',Arial,sans-serif;background:#F9FAFB;padding:24px;color:#111827;">
<div style="max-width:600px;margin:0 auto;background:#FFFFFF;border-radius:12px;overflow:hidden;
            box-shadow:0 8px 24px rgba(0,0,0,0.06);border:1px solid #E5E7EB;">

  <!-- Header band -->
  <div style="background:linear-gradient(135deg,#1E3A5F 0%,#4F46E5 100%);color:#FFF;padding:24px 28px;">
    <div style="font-size:0.72rem;letter-spacing:0.15em;color:#C9A84C;font-weight:700;">QUOTATION</div>
    <div style="font-size:1.4rem;font-weight:800;margin-top:4px;">{COMPANY['short']}</div>
    <div style="font-size:0.75rem;color:#CBD5E1;margin-top:2px;">Enterprise Bitumen Desk · {COMPANY['city']}</div>
  </div>

  <!-- Date + Recipient -->
  <div style="padding:20px 28px;border-bottom:1px solid #E5E7EB;">
    <div style="font-size:0.72rem;color:#6B7280;text-transform:uppercase;letter-spacing:0.08em;">Date</div>
    <div style="font-weight:600;">{_today_long()}</div>
    <div style="margin-top:12px;font-size:0.72rem;color:#6B7280;text-transform:uppercase;letter-spacing:0.08em;">To</div>
    <div style="font-weight:700;font-size:1.05rem;">{customer_name}</div>
    <div style="color:#6B7280;font-size:0.9rem;">📍 {city}{' · 🏭 ' + source if source else ''}</div>
  </div>

  <!-- Quote table -->
  <div style="padding:20px 28px;">
    <table style="width:100%;border-collapse:collapse;font-size:0.95rem;">
      <tr><td style="padding:10px 0;color:#6B7280;">Grade</td>
          <td style="padding:10px 0;text-align:right;font-weight:700;">{grade}</td></tr>
      <tr><td style="padding:10px 0;color:#6B7280;border-top:1px solid #F3F4F6;">Quantity</td>
          <td style="padding:10px 0;text-align:right;font-weight:700;border-top:1px solid #F3F4F6;">{qty_mt:,.0f} MT</td></tr>
      <tr><td style="padding:10px 0;color:#6B7280;border-top:1px solid #F3F4F6;">Rate per MT</td>
          <td style="padding:10px 0;text-align:right;font-weight:700;border-top:1px solid #F3F4F6;">{fmt_inr(price_per_mt)}</td></tr>
      <tr><td style="padding:14px 0 10px;font-weight:800;color:#1E3A5F;border-top:2px solid #1E3A5F;">Total Value</td>
          <td style="padding:14px 0 10px;text-align:right;font-weight:800;color:#1E3A5F;border-top:2px solid #1E3A5F;font-size:1.15rem;">{fmt_inr(total)}</td></tr>
    </table>
  </div>

  <!-- Terms band -->
  <div style="background:#F9FAFB;padding:16px 28px;border-top:1px solid #E5E7EB;border-bottom:1px solid #E5E7EB;
              font-size:0.82rem;color:#4B5563;line-height:1.7;">
    🚚 Dispatch within 48 hrs · ⏳ Validity 24 hrs · 💳 100% Advance · GST 18% (HSN {COMPANY['hsn']})
  </div>

  <!-- Bank block -->
  <div style="padding:16px 28px;border-bottom:1px solid #E5E7EB;">
    <div style="font-size:0.72rem;color:#6B7280;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:6px;">Payment To</div>
    <div style="font-size:0.95rem;line-height:1.6;">
      <b>{COMPANY['bank']}</b><br>
      A/C: <code style="background:#F3F4F6;padding:2px 6px;border-radius:4px;">{COMPANY['bank_ac']}</code>
      · IFSC: <code style="background:#F3F4F6;padding:2px 6px;border-radius:4px;">{COMPANY['bank_ifsc']}</code><br>
      <span style="color:#6B7280;font-size:0.85rem;">GST: {COMPANY['gst']}</span>
    </div>
  </div>

  <!-- CTA -->
  <div style="padding:20px 28px;text-align:center;background:#FFFBEB;">
    <div style="font-size:0.95rem;color:#92400E;">
      ✅ <b>Reply to lock this price</b> · 📞 {COMPANY['phone']}
    </div>
  </div>

  <!-- Footer -->
  <div style="padding:14px 28px;font-size:0.7rem;color:#9CA3AF;background:#F9FAFB;">
    {COMPANY['name']} · {COMPANY['city']}<br>
    CIN: {COMPANY['cin']} · PAN: {COMPANY['pan']}
  </div>
</div></body></html>
"""

    body_text = (
        f"{COMPANY['short']} — Bitumen Quotation\n"
        f"Date: {_today_long()}\n\n"
        f"To: {customer_name}\nLocation: {city}\n\n"
        f"Grade: {grade}\nQuantity: {qty_mt:,.0f} MT\n"
        f"Rate: {fmt_inr(price_per_mt)}/MT\n"
        f"Total: {fmt_inr(total)}\n\n"
        f"Terms: Dispatch 48 hrs, Validity 24 hrs, 100% advance, GST 18%\n\n"
        f"Bank: {COMPANY['bank']} · A/C {COMPANY['bank_ac']} · IFSC {COMPANY['bank_ifsc']}\n\n"
        f"Reply to lock this price · {COMPANY['phone']}\n"
        f"— {COMPANY['short']}"
    )

    return {"subject": subject, "body_html": body_html, "body_text": body_text}


# ═══════════════════════════════════════════════════════════════════════════
# Excel — openpyxl styled
# ═══════════════════════════════════════════════════════════════════════════

def build_styled_excel(sheets: dict[str, list[dict]],
                        title: str = "PPS Anantam Report") -> bytes:
    """
    Build a branded XLSX with one sheet per dict entry.
    Each value is a list of dict rows.

    Features: header band (company name + title), frozen header row,
    auto-width, currency formatting for ₹ columns, alternating row colors,
    footer with generated timestamp.

    Returns empty bytes if openpyxl is unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border,
                                      Side, NamedStyle)
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback: return naive XLSX via pandas if available
        try:
            import pandas as pd
            buf = _io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                for name, rows in sheets.items():
                    pd.DataFrame(rows).to_excel(writer, sheet_name=name[:31],
                                                 index=False)
            return buf.getvalue()
        except Exception:
            return b""

    wb = Workbook()
    wb.remove(wb.active)

    header_font    = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill    = PatternFill("solid", fgColor="1E3A5F")
    brand_font     = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
    sub_font       = Font(name="Calibri", size=9, italic=True, color="6B7280")
    alt_fill       = PatternFill("solid", fgColor="F9FAFB")
    thin_border    = Border(bottom=Side(style="thin", color="E5E7EB"))
    footer_font    = Font(name="Calibri", size=8, italic=True, color="9CA3AF")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        if not rows:
            ws["A1"] = f"{title} — {sheet_name}"
            ws["A1"].font = brand_font
            ws["A3"] = "(No data)"
            continue

        # Branded header block (row 1-3)
        ws["A1"] = COMPANY["short"]
        ws["A1"].font = brand_font
        ws["A2"] = f"{title} — {sheet_name}"
        ws["A2"].font = Font(name="Calibri", size=10, bold=True, color="4F46E5")
        ws["A3"] = (f"{COMPANY['city']} · GST {COMPANY['gst']} · "
                    f"Generated {_dt.datetime.now().strftime('%d %b %Y %H:%M IST')}")
        ws["A3"].font = sub_font

        # Data headers on row 5
        headers = list(rows[0].keys())
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=j, value=h.replace("_", " ").title())
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

        # Data rows from row 6
        for i, row in enumerate(rows, 6):
            for j, h in enumerate(headers, 1):
                c = ws.cell(row=i, column=j, value=row.get(h))
                if (i - 6) % 2 == 1:
                    c.fill = alt_fill
                # INR formatting for amount-like columns
                if any(k in h.lower() for k in ("price", "amount", "total",
                                                 "value", "rate", "cost")):
                    try:
                        c.value = float(row.get(h))
                        c.number_format = '"\u20b9"#,##,##0.00'
                    except (TypeError, ValueError):
                        pass
                c.border = thin_border

        # Freeze header
        ws.freeze_panes = "A6"

        # Column widths based on max content
        for j, h in enumerate(headers, 1):
            max_len = max(
                [len(str(h))] +
                [len(str(r.get(h, ""))) for r in rows]
            )
            ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 12), 40)

        # Footer row
        footer_row = len(rows) + 7
        ws.cell(row=footer_row, column=1,
                value=f"{COMPANY['name']} · CIN {COMPANY['cin']} · Confidential"
               ).font = footer_font

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# PDF — premium ReportLab quote
# ═══════════════════════════════════════════════════════════════════════════

def build_quote_pdf(customer_name: str, city: str, grade: str,
                    qty_mt: float, price_per_mt: float,
                    source: str = "",
                    quote_no: str = "",
                    terms: list[str] | None = None) -> bytes:
    """
    Build a premium A4 quote PDF with:
      - Colored header band (navy/indigo)
      - Quote number + date + customer block
      - Styled items table with total row highlighted
      - Terms list
      - Bank details card
      - Signature block
      - Footer with CIN/GST

    Returns empty bytes if reportlab is unavailable.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle, ListFlowable, ListItem)
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return b""

    # ── Font registration for ₹ glyph support ───────────────────────
    # Default Helvetica lacks ₹ glyph → renders as tofu. Try common
    # system Unicode TTFs; fall back to "Rs." prefix if none register.
    _font_name = "Helvetica"
    _font_bold = "Helvetica-Bold"
    _rupee     = "Rs."
    for candidate, bold_candidate, paths in [
        ("DejaVuSans", "DejaVuSans-Bold", [
            "C:/Windows/Fonts/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/Library/Fonts/DejaVuSans.ttf",
        ]),
        ("NotoSans", "NotoSans-Bold", [
            "C:/Windows/Fonts/NotoSans-Regular.ttf",
            "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        ]),
        ("Arial", "Arial-Bold", [
            "C:/Windows/Fonts/arial.ttf",
            "/Library/Fonts/Arial.ttf",
        ]),
    ]:
        for path in paths:
            try:
                import os as _os
                if _os.path.exists(path):
                    pdfmetrics.registerFont(TTFont(candidate, path))
                    # Try to register matching bold (best-effort)
                    bold_path = path.replace("Regular", "Bold").replace(".ttf", "-Bold.ttf")
                    alt_bold = path.replace(".ttf", "-Bold.ttf")
                    for bp in (bold_path, alt_bold,
                               "C:/Windows/Fonts/arialbd.ttf" if candidate == "Arial" else None):
                        if bp and _os.path.exists(bp):
                            pdfmetrics.registerFont(TTFont(bold_candidate, bp))
                            _font_bold = bold_candidate
                            break
                    else:
                        _font_bold = candidate  # use regular for bold if no bold
                    _font_name = candidate
                    _rupee = "\u20b9"
                    break
            except Exception:
                continue
        if _font_name != "Helvetica":
            break

    if terms is None:
        terms = [
            "Rates are ex-terminal / ex-warehouse. GST 18% (HSN 27132000) extra.",
            "Payment: 100% advance against pro-forma invoice.",
            "Validity: 24 hours from quotation date & time.",
            "Dispatch: Within 48 hours of confirmed payment.",
            "Quality: As per IS 73:2018 standard for bitumen.",
            "Delivery at buyer's risk. Transit insurance not included.",
            "Any dispute subject to Vadodara, Gujarat jurisdiction.",
            "Prices are subject to revision without notice until confirmation.",
        ]

    if not quote_no:
        quote_no = _dt.datetime.now().strftime("Q%d%m%Y-%H%M")

    total_value = qty_mt * price_per_mt

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15 * mm, rightMargin=15 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm,
                             title=f"Quotation {quote_no} — {customer_name}",
                             author=COMPANY["name"])

    styles = getSampleStyleSheet()
    brand_title = ParagraphStyle("brand_title", parent=styles["Heading1"],
                                  fontName=_font_bold,
                                  fontSize=18, textColor=colors.white,
                                  spaceAfter=0, leading=22)
    brand_sub   = ParagraphStyle("brand_sub", parent=styles["Normal"],
                                  fontName=_font_bold,
                                  fontSize=8, textColor=colors.HexColor("#C9A84C"),
                                  spaceBefore=2, leading=11)
    brand_loc   = ParagraphStyle("brand_loc", parent=styles["Normal"],
                                  fontName=_font_name,
                                  fontSize=8, textColor=colors.HexColor("#CBD5E1"),
                                  leading=11)
    section_h   = ParagraphStyle("section_h", parent=styles["Heading3"],
                                  fontName=_font_bold,
                                  fontSize=10, textColor=colors.HexColor("#1E3A5F"),
                                  spaceAfter=4, spaceBefore=10, leading=14)
    normal      = ParagraphStyle("normal_p", parent=styles["Normal"],
                                  fontName=_font_name,
                                  fontSize=9, leading=14)
    term_p      = ParagraphStyle("term_p", parent=styles["Normal"],
                                  fontName=_font_name,
                                  fontSize=9, leading=14,
                                  leftIndent=14, spaceAfter=4,
                                  bulletIndent=0)
    footer      = ParagraphStyle("footer_p", parent=styles["Normal"],
                                  fontName=_font_name,
                                  fontSize=7, textColor=colors.HexColor("#9CA3AF"),
                                  alignment=TA_CENTER, leading=9)

    story = []

    # ── Header band with logo (2-col) ───────────────────────────────
    import os as _os
    _logo_path = _os.path.join(_os.path.dirname(__file__), COMPANY["logo"])

    # Left: logo if exists, else spacer
    if _os.path.exists(_logo_path):
        try:
            from reportlab.platypus import Image as _RLImage
            logo_obj = _RLImage(_logo_path, width=22 * mm, height=22 * mm)
        except Exception:
            logo_obj = ""
    else:
        logo_obj = ""

    # Right side: company title + tagline + location
    right_block = [
        Paragraph("<b>QUOTATION</b>", brand_sub),
        Paragraph(COMPANY["short"], brand_title),
        Paragraph(f"<font color='#C9A84C'><i>{COMPANY['tagline']}</i></font>",
                  ParagraphStyle("tagline", parent=styles["Normal"],
                                  fontName=_font_name, fontSize=7.5,
                                  textColor=colors.HexColor("#C9A84C"),
                                  leading=10)),
        Paragraph(f"Enterprise Bitumen Desk &middot; {COMPANY['city']}", brand_loc),
    ]
    right_tbl = Table([[p] for p in right_block], colWidths=[152 * mm])
    right_tbl.setStyle(TableStyle([
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 1),
    ]))

    hdr = Table([[logo_obj, right_tbl]], colWidths=[26 * mm, 154 * mm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor(BRAND["navy"])),
        ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
        ("TOPPADDING",    (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(hdr)

    # Thin gold divider
    gold_div = Table([[""]], colWidths=[180 * mm], rowHeights=[2.5])
    gold_div.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["gold"])),
    ]))
    story.append(gold_div)
    story.append(Spacer(1, 8))

    # ── Quote meta + Customer block (2-col) ─────────────────────────
    meta_tbl = Table([[
        Paragraph(f"<b>Quotation #:</b> {quote_no}<br/>"
                  f"<b>Date:</b> {_today_long()} · {_now_time()}",
                  normal),
        Paragraph(f"<b>To:</b> {customer_name}<br/>"
                  f"<b>Location:</b> {city}"
                  + (f"<br/><b>Source:</b> {source}" if source else ""),
                  normal),
    ]], colWidths=[85 * mm, 95 * mm])
    meta_tbl.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 12))

    # ── Items table ─────────────────────────────────────────────────
    def _fmt_amt(n):
        s = fmt_inr(n)
        # Strip ₹ for table cells (header shows the symbol once)
        return s.lstrip("\u20b9")

    items_data = [
        ["#", "Description",                             "Qty (MT)",         f"Rate ({_rupee}/MT)",    f"Value ({_rupee})"],
        ["1", f"Bitumen {grade} — HSN {COMPANY['hsn']}", f"{qty_mt:,.0f}",   _fmt_amt(price_per_mt),    _fmt_amt(total_value)],
        ["",  "GST @ 18%",                                "",                 "",                        _fmt_amt(total_value * 0.18)],
        ["",  "Grand Total (incl. GST)",                  "",                 "",                        _fmt_amt(total_value * 1.18)],
    ]
    items = Table(items_data, colWidths=[12 * mm, 80 * mm, 26 * mm, 28 * mm, 34 * mm])
    items.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), _font_bold),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        # Body font
        ("FONTNAME",   (0, 1), (-1, -1), _font_name),
        ("ALIGN",      (2, 1), (-1, -1), "RIGHT"),
        # GST row
        ("TEXTCOLOR",  (0, 2), (-1, 2), colors.HexColor("#4B5563")),
        # Grand total
        ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#FFFBEB")),
        ("FONTNAME",   (0, 3), (-1, 3), _font_bold),
        ("TEXTCOLOR",  (0, 3), (-1, 3), colors.HexColor("#92400E")),
        ("FONTSIZE",   (0, 3), (-1, 3), 11),
        # Grid
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(items)

    # ── Total in Words (Indian invoicing standard) ──────────────────
    grand_total = total_value * 1.18
    words = _num_to_words_indian(int(round(grand_total)))
    if words:
        tw_style = ParagraphStyle("tw", parent=styles["Normal"],
                                   fontName=_font_name,
                                   fontSize=8.5, leading=11,
                                   textColor=colors.HexColor(BRAND["grey_700"]))
        tw_tbl = Table([[
            Paragraph(f"<b>Amount in words:</b> Rupees {words} Only",
                      tw_style)
        ]], colWidths=[180 * mm])
        tw_tbl.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, -1), colors.HexColor(BRAND["grey_50"])),
            ("LINEBELOW",    (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
            ("LINEABOVE",    (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
            ("LEFTPADDING",  (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING",   (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ]))
        story.append(tw_tbl)
    story.append(Spacer(1, 12))

    # ── Terms (each as its own paragraph, numbered) ─────────────────
    story.append(Paragraph("Terms &amp; Conditions", section_h))
    for i, t in enumerate(terms, 1):
        story.append(Paragraph(
            f'<font color="#4F46E5"><b>{i}.</b></font> &nbsp;{t}',
            term_p))
    story.append(Spacer(1, 8))

    # ── Thank-you band (gentle green) ───────────────────────────────
    ty_style = ParagraphStyle("ty", parent=styles["Normal"],
                               fontName=_font_bold, fontSize=10,
                               textColor=colors.HexColor(BRAND["green_fg"]),
                               alignment=TA_CENTER, leading=13)
    ty_sub = ParagraphStyle("tysub", parent=styles["Normal"],
                             fontName=_font_name, fontSize=8,
                             textColor=colors.HexColor(BRAND["green_fg"]),
                             alignment=TA_CENTER, leading=11)
    ty_tbl = Table([[
        Paragraph("THANK YOU FOR YOUR BUSINESS", ty_style),
        ],[
        Paragraph(f"Reply <b>CONFIRM</b> to lock this price &middot; Call {COMPANY['phone']} &middot; {COMPANY['email']}",
                  ty_sub),
    ]], colWidths=[180 * mm])
    ty_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), colors.HexColor(BRAND["green_soft"])),
        ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#A7F3D0")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 14),
        ("RIGHTPADDING", (0, 0), (-1, -1), 14),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
    ]))
    story.append(ty_tbl)
    story.append(Spacer(1, 10))

    # ── Bank + Signature (2-col) ─────────────────────────────────────
    bank_p = Paragraph(
        f"<b>Bank Details</b><br/>"
        f"{COMPANY['bank']}<br/>"
        f"A/C: <b>{COMPANY['bank_ac']}</b><br/>"
        f"IFSC: <b>{COMPANY['bank_ifsc']}</b><br/>"
        f"GST: {COMPANY['gst']}<br/>"
        f"Total Payable: <b>{_rupee}{_fmt_amt(total_value * 1.18)}</b>",
        normal)
    sig_p = Paragraph(
        f"<b>For {COMPANY['name']}</b><br/><br/><br/>"
        f"_______________________<br/>"
        f"Authorized Signatory<br/>"
        f"<font color='#6B7280' size='8'>{COMPANY['owner']} &middot; {COMPANY['phone']}</font>",
        normal)
    bs = Table([[bank_p, sig_p]], colWidths=[90 * mm, 90 * mm])
    bs.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#F9FAFB")),
        ("BOX",        (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("LINEBETWEEN",(0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("LEFTPADDING",(0, 0), (-1, -1), 10),
        ("RIGHTPADDING",(0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 10),
    ]))
    story.append(bs)
    story.append(Spacer(1, 10))

    # ── Footer ───────────────────────────────────────────────────────
    story.append(Paragraph(
        f"{COMPANY['name']} &middot; CIN {COMPANY['cin']} &middot; "
        f"PAN {COMPANY['pan']} &middot; {COMPANY['city']} &middot; "
        f"Confidential — For {customer_name} only",
        footer))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# Shared branded header helper
# ═══════════════════════════════════════════════════════════════════════════

def _branded_header(doc_type_label: str, _font_name: str, _font_bold: str,
                     styles, mm, Table, TableStyle, Paragraph, Spacer, colors,
                     ParagraphStyle):
    """Shared header band with logo + title + tagline. Returns list of flowables."""
    import os as _os
    from reportlab.platypus import Image as _RLImage

    brand_title = ParagraphStyle("brand_title_" + doc_type_label,
                                  parent=styles["Heading1"],
                                  fontName=_font_bold, fontSize=18,
                                  textColor=colors.white, leading=22)
    brand_sub   = ParagraphStyle("brand_sub_" + doc_type_label,
                                  parent=styles["Normal"], fontName=_font_bold,
                                  fontSize=8, textColor=colors.HexColor(BRAND["gold"]),
                                  spaceBefore=2, leading=11)
    brand_tag   = ParagraphStyle("brand_tag_" + doc_type_label,
                                  parent=styles["Normal"], fontName=_font_name,
                                  fontSize=7.5, textColor=colors.HexColor(BRAND["gold"]),
                                  leading=10)
    brand_loc   = ParagraphStyle("brand_loc_" + doc_type_label,
                                  parent=styles["Normal"], fontName=_font_name,
                                  fontSize=8, textColor=colors.HexColor("#CBD5E1"),
                                  leading=11)

    logo_path = _os.path.join(_os.path.dirname(__file__), COMPANY["logo"])
    logo_obj = ""
    if _os.path.exists(logo_path):
        try:
            logo_obj = _RLImage(logo_path, width=22 * mm, height=22 * mm)
        except Exception:
            pass

    right_block = [
        Paragraph(f"<b>{doc_type_label}</b>", brand_sub),
        Paragraph(COMPANY["short"], brand_title),
        Paragraph(f"<i>{COMPANY['tagline']}</i>", brand_tag),
        Paragraph(f"Enterprise Bitumen Desk &middot; {COMPANY['city']}", brand_loc),
    ]
    right_tbl = Table([[p] for p in right_block], colWidths=[152 * mm])
    right_tbl.setStyle(TableStyle([
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 1),
    ]))
    hdr = Table([[logo_obj, right_tbl]], colWidths=[26 * mm, 154 * mm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), colors.HexColor(BRAND["navy"])),
        ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 14),
        ("TOPPADDING",    (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    gold_div = Table([[""]], colWidths=[180 * mm], rowHeights=[2.5])
    gold_div.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["gold"])),
    ]))
    return [hdr, gold_div, Spacer(1, 8)]


# ═══════════════════════════════════════════════════════════════════════════
# TAX INVOICE
# ═══════════════════════════════════════════════════════════════════════════

def build_invoice_pdf(customer_name: str, customer_gst: str,
                      customer_address: str, customer_state: str,
                      invoice_no: str, invoice_date: str,
                      items: list[dict],  # [{'desc','qty','rate'}, ...]
                      place_of_supply: str = "") -> bytes:
    """
    Premium Tax Invoice PDF.
    items = list of dicts with desc/qty/rate (₹ per MT).
    IGST if inter-state, else CGST+SGST.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle)
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return b""

    # Font registration (same as quote)
    _font_name = "Helvetica"
    _font_bold = "Helvetica-Bold"
    _rupee = "Rs."
    for candidate, bold_candidate, paths in [
        ("DejaVuSans", "DejaVuSans-Bold",
         ["C:/Windows/Fonts/DejaVuSans.ttf",
          "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]),
        ("Arial", "Arial-Bold",
         ["C:/Windows/Fonts/arial.ttf", "/Library/Fonts/Arial.ttf"]),
    ]:
        for p in paths:
            try:
                import os as _os
                if _os.path.exists(p):
                    pdfmetrics.registerFont(TTFont(candidate, p))
                    bp = p.replace(".ttf", "-Bold.ttf")
                    if candidate == "Arial":
                        bp = "C:/Windows/Fonts/arialbd.ttf"
                    if _os.path.exists(bp):
                        pdfmetrics.registerFont(TTFont(bold_candidate, bp))
                        _font_bold = bold_candidate
                    else:
                        _font_bold = candidate
                    _font_name = candidate
                    _rupee = "\u20b9"
                    break
            except Exception:
                continue
        if _font_name != "Helvetica":
            break

    is_inter_state = customer_state and "Gujarat" not in customer_state

    # Compute totals
    taxable = sum(it["qty"] * it["rate"] for it in items)
    if is_inter_state:
        igst = taxable * 0.18
        cgst = sgst = 0
    else:
        cgst = sgst = taxable * 0.09
        igst = 0
    grand = taxable + cgst + sgst + igst

    def _fmt_amt(n):
        return fmt_inr(n).lstrip("\u20b9")

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15 * mm, rightMargin=15 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm,
                             title=f"Invoice {invoice_no}",
                             author=COMPANY["name"])
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("n", parent=styles["Normal"], fontName=_font_name,
                             fontSize=9, leading=13)
    small  = ParagraphStyle("s", parent=styles["Normal"], fontName=_font_name,
                             fontSize=7.5, textColor=colors.HexColor(BRAND["grey_500"]),
                             leading=10)
    section_h = ParagraphStyle("sh", parent=styles["Heading3"], fontName=_font_bold,
                                fontSize=10, textColor=colors.HexColor(BRAND["navy_light"]),
                                spaceAfter=4, spaceBefore=10, leading=14)
    footer = ParagraphStyle("f", parent=styles["Normal"], fontName=_font_name,
                             fontSize=7, textColor=colors.HexColor("#9CA3AF"),
                             alignment=TA_CENTER, leading=9)

    story = []
    story.extend(_branded_header("TAX INVOICE", _font_name, _font_bold, styles,
                                  mm, Table, TableStyle, Paragraph, Spacer,
                                  colors, ParagraphStyle))

    # Invoice meta + parties
    meta = Table([[
        Paragraph(f"<b>Invoice #:</b> {invoice_no}<br/>"
                  f"<b>Date:</b> {invoice_date}<br/>"
                  f"<b>Place of Supply:</b> {place_of_supply or customer_state}<br/>"
                  f"<b>Reverse Charge:</b> No", normal),
        Paragraph(f"<b>Bill To:</b><br/>"
                  f"<b>{customer_name}</b><br/>"
                  f"{customer_address}<br/>"
                  f"GSTIN: {customer_gst or 'Unregistered'}<br/>"
                  f"State: {customer_state}", normal),
    ]], colWidths=[85 * mm, 95 * mm])
    meta.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta)
    story.append(Spacer(1, 12))

    # Items table
    rows = [["#", "Description", "HSN", "Qty (MT)", f"Rate ({_rupee})", f"Amount ({_rupee})"]]
    for i, it in enumerate(items, 1):
        amt = it["qty"] * it["rate"]
        rows.append([str(i), it.get("desc", "Bitumen"), COMPANY["hsn"],
                     f"{it['qty']:,.0f}", _fmt_amt(it["rate"]), _fmt_amt(amt)])
    rows.append(["", "Taxable Value", "", "", "", _fmt_amt(taxable)])
    if is_inter_state:
        rows.append(["", "IGST @ 18%", "", "", "", _fmt_amt(igst)])
    else:
        rows.append(["", "CGST @ 9%", "", "", "", _fmt_amt(cgst)])
        rows.append(["", "SGST @ 9%", "", "", "", _fmt_amt(sgst)])
    rows.append(["", "Grand Total", "", "", "", _fmt_amt(grand)])

    items_tbl = Table(rows,
                      colWidths=[10 * mm, 62 * mm, 22 * mm, 22 * mm, 28 * mm, 36 * mm])
    last_row = len(rows) - 1
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND["navy"])),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), _font_bold),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME",   (0, 1), (-1, -1), _font_name),
        ("ALIGN",      (3, 1), (-1, -1), "RIGHT"),
        # Grand total row
        ("BACKGROUND", (0, last_row), (-1, last_row), colors.HexColor(BRAND["amber_bg"])),
        ("FONTNAME",   (0, last_row), (-1, last_row), _font_bold),
        ("TEXTCOLOR",  (0, last_row), (-1, last_row), colors.HexColor(BRAND["amber_fg"])),
        ("FONTSIZE",   (0, last_row), (-1, last_row), 11),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(items_tbl)

    # Amount in words
    words = _num_to_words_indian(int(round(grand)))
    if words:
        tw = Table([[Paragraph(f"<b>Amount in words:</b> Rupees {words} Only",
                                ParagraphStyle("tw", parent=styles["Normal"],
                                                fontName=_font_name, fontSize=8.5,
                                                leading=11))]],
                   colWidths=[180 * mm])
        tw.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor(BRAND["grey_50"])),
            ("BOX",(0,0),(-1,-1),0.4,colors.HexColor(BRAND["grey_200"])),
            ("LEFTPADDING",(0,0),(-1,-1),10),
            ("TOPPADDING",(0,0),(-1,-1),6),
            ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ]))
        story.append(tw)
    story.append(Spacer(1, 10))

    # Declarations
    story.append(Paragraph("Declarations", section_h))
    decl = [
        "We declare that this invoice shows the actual price of goods described and all particulars are true and correct.",
        "Bitumen conforms to IS 73:2018 specification for VG grade bitumen.",
        "All disputes subject to Vadodara, Gujarat jurisdiction only.",
        "Payment due per agreed terms; interest @ 18% p.a. on overdue balances.",
    ]
    for i, d in enumerate(decl, 1):
        story.append(Paragraph(f"<font color='{BRAND['navy_light']}'><b>{i}.</b></font> &nbsp;{d}",
                                ParagraphStyle("d", parent=styles["Normal"],
                                                fontName=_font_name, fontSize=9,
                                                leading=13, leftIndent=14, spaceAfter=4)))
    story.append(Spacer(1, 10))

    # Bank + signature
    bank_p = Paragraph(
        f"<b>Payment Details</b><br/>"
        f"{COMPANY['bank']}<br/>"
        f"A/C: <b>{COMPANY['bank_ac']}</b><br/>"
        f"IFSC: <b>{COMPANY['bank_ifsc']}</b><br/>"
        f"GST: {COMPANY['gst']}<br/>"
        f"<b>Amount Due: {_rupee}{_fmt_amt(grand)}</b>", normal)
    sig_p = Paragraph(
        f"<b>For {COMPANY['name']}</b><br/><br/><br/>"
        f"_______________________<br/>"
        f"Authorized Signatory<br/>"
        f"<font color='{BRAND['grey_500']}' size='8'>{COMPANY['owner']} &middot; {COMPANY['phone']}</font>",
        normal)
    bs = Table([[bank_p, sig_p]], colWidths=[90 * mm, 90 * mm])
    bs.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(BRAND["grey_50"])),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(bs)
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"{COMPANY['name']} &middot; CIN {COMPANY['cin']} &middot; "
        f"PAN {COMPANY['pan']} &middot; {COMPANY['city']} &middot; "
        f"Original Copy for Recipient", footer))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# PURCHASE ORDER
# ═══════════════════════════════════════════════════════════════════════════

def build_purchase_order_pdf(supplier_name: str, supplier_address: str,
                             supplier_gst: str,
                             po_no: str, po_date: str,
                             delivery_date: str, delivery_address: str,
                             items: list[dict]) -> bytes:
    """Premium Purchase Order PDF (bitumen buyer side)."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle)
        from reportlab.lib.enums import TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return b""

    _font_name = "Helvetica"
    _font_bold = "Helvetica-Bold"
    _rupee = "Rs."
    for p in ["C:/Windows/Fonts/arial.ttf",
              "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]:
        import os as _os
        if _os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("MainFont", p))
                _font_name = "MainFont"
                bp = "C:/Windows/Fonts/arialbd.ttf" if "arial" in p else p.replace(".ttf", "-Bold.ttf")
                if _os.path.exists(bp):
                    pdfmetrics.registerFont(TTFont("MainBold", bp))
                    _font_bold = "MainBold"
                else:
                    _font_bold = "MainFont"
                _rupee = "\u20b9"
                break
            except Exception:
                continue

    def _fmt_amt(n):
        return fmt_inr(n).lstrip("\u20b9")

    taxable = sum(it["qty"] * it["rate"] for it in items)
    gst = taxable * 0.18
    grand = taxable + gst

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15 * mm, rightMargin=15 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm,
                             title=f"Purchase Order {po_no}",
                             author=COMPANY["name"])
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("n", parent=styles["Normal"], fontName=_font_name,
                             fontSize=9, leading=13)
    section_h = ParagraphStyle("sh", parent=styles["Heading3"], fontName=_font_bold,
                                fontSize=10, textColor=colors.HexColor(BRAND["navy_light"]),
                                spaceAfter=4, spaceBefore=10, leading=14)
    footer = ParagraphStyle("f", parent=styles["Normal"], fontName=_font_name,
                             fontSize=7, textColor=colors.HexColor("#9CA3AF"),
                             alignment=TA_CENTER, leading=9)

    story = []
    story.extend(_branded_header("PURCHASE ORDER", _font_name, _font_bold, styles,
                                  mm, Table, TableStyle, Paragraph, Spacer,
                                  colors, ParagraphStyle))

    # Meta block
    meta = Table([[
        Paragraph(f"<b>PO #:</b> {po_no}<br/>"
                  f"<b>PO Date:</b> {po_date}<br/>"
                  f"<b>Required By:</b> {delivery_date}<br/>"
                  f"<b>Buyer GST:</b> {COMPANY['gst']}", normal),
        Paragraph(f"<b>Supplier:</b><br/><b>{supplier_name}</b><br/>"
                  f"{supplier_address}<br/>"
                  f"GSTIN: {supplier_gst or 'Unregistered'}", normal),
    ]], colWidths=[85 * mm, 95 * mm])
    meta.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta)
    story.append(Spacer(1, 10))

    # Delivery address
    da = Table([[Paragraph(
        f"<b>Ship To:</b> {delivery_address}", normal)]],
        colWidths=[180 * mm])
    da.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["grey_50"])),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(da)
    story.append(Spacer(1, 12))

    # Items table
    rows = [["#", "Description", "HSN", "Qty (MT)", f"Rate ({_rupee}/MT)", f"Amount ({_rupee})"]]
    for i, it in enumerate(items, 1):
        rows.append([str(i), it.get("desc", "Bitumen VG30"), COMPANY["hsn"],
                     f"{it['qty']:,.0f}", _fmt_amt(it["rate"]),
                     _fmt_amt(it["qty"] * it["rate"])])
    rows.append(["", "Sub-total", "", "", "", _fmt_amt(taxable)])
    rows.append(["", "GST @ 18%", "", "", "", _fmt_amt(gst)])
    rows.append(["", "Order Value", "", "", "", _fmt_amt(grand)])
    items_tbl = Table(rows,
                      colWidths=[10 * mm, 62 * mm, 22 * mm, 22 * mm, 28 * mm, 36 * mm])
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND["navy"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), _font_bold),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), _font_name),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(BRAND["amber_bg"])),
        ("FONTNAME", (0, -1), (-1, -1), _font_bold),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor(BRAND["amber_fg"])),
        ("FONTSIZE", (0, -1), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 12))

    # Terms
    story.append(Paragraph("Order Terms", section_h))
    terms = [
        f"Delivery required by {delivery_date} at specified address.",
        "Goods must conform to IS 73:2018 bitumen specification.",
        "Quality test certificate to accompany each tanker.",
        "Payment: 50% advance, 50% on receipt + quality acceptance.",
        "Any delay beyond agreed date attracts 2% penalty per week.",
        "All disputes subject to Vadodara, Gujarat jurisdiction.",
    ]
    for i, t in enumerate(terms, 1):
        story.append(Paragraph(
            f"<font color='{BRAND['navy_light']}'><b>{i}.</b></font> &nbsp;{t}",
            ParagraphStyle("t", parent=styles["Normal"], fontName=_font_name,
                            fontSize=9, leading=13, leftIndent=14, spaceAfter=3)))
    story.append(Spacer(1, 10))

    # Authorization
    auth = Table([[
        Paragraph(f"<b>For {COMPANY['name']} (Buyer)</b><br/><br/><br/>"
                  f"_______________________<br/>"
                  f"Authorized Signatory<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>{COMPANY['owner']} &middot; {COMPANY['phone']}</font>",
                  normal),
        Paragraph(f"<b>Acknowledged by Supplier</b><br/><br/><br/>"
                  f"_______________________<br/>"
                  f"Signature &middot; Stamp<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>{supplier_name}</font>",
                  normal),
    ]], colWidths=[90 * mm, 90 * mm])
    auth.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(auth)
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"{COMPANY['name']} &middot; GST {COMPANY['gst']} &middot; "
        f"{COMPANY['city']} &middot; PO confidential — For {supplier_name} only",
        footer))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# DIRECTOR BRIEF (executive report)
# ═══════════════════════════════════════════════════════════════════════════

def build_director_brief_pdf(date_str: str, kpis: list[dict],
                              sections: list[dict],
                              overall_status: str = "GOOD",
                              top_action: str = "",
                              mood: str = "") -> bytes:
    """
    Premium executive brief PDF — v2 executive dashboard style.

    kpis     = [{'label','value','delta','color','direction'}]  # direction: up/down/flat
    sections = [{'title','bullets':[...], 'accent': '#color', 'icon': '●'}]
    overall_status = "GOOD" / "WATCH" / "ALERT"
    top_action = one-line call-to-action shown in amber band
    mood = optional subtitle / day-mood line
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle)
        from reportlab.lib.enums import TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return b""

    _font_name = "Helvetica"
    _font_bold = "Helvetica-Bold"
    for p in ["C:/Windows/Fonts/arial.ttf",
              "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]:
        import os as _os
        if _os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("MainFontDB", p))
                _font_name = "MainFontDB"
                bp = "C:/Windows/Fonts/arialbd.ttf" if "arial" in p else p.replace(".ttf", "-Bold.ttf")
                if _os.path.exists(bp):
                    pdfmetrics.registerFont(TTFont("MainBoldDB", bp))
                    _font_bold = "MainBoldDB"
                else:
                    _font_bold = "MainFontDB"
                break
            except Exception:
                continue

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15 * mm, rightMargin=15 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm,
                             title="Director Brief",
                             author=COMPANY["name"])
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("n", parent=styles["Normal"], fontName=_font_name,
                             fontSize=9.5, leading=14)
    bullet = ParagraphStyle("b", parent=styles["Normal"], fontName=_font_name,
                             fontSize=9.5, leading=14, leftIndent=12, spaceAfter=4)
    section_h = ParagraphStyle("sh", parent=styles["Heading2"], fontName=_font_bold,
                                fontSize=11.5, textColor=colors.HexColor(BRAND["navy_light"]),
                                spaceAfter=6, spaceBefore=12, leading=14)
    kpi_label = ParagraphStyle("kl", parent=styles["Normal"], fontName=_font_name,
                                fontSize=7.5, textColor=colors.HexColor(BRAND["grey_500"]),
                                alignment=TA_CENTER, leading=10)
    kpi_val = ParagraphStyle("kv", parent=styles["Normal"], fontName=_font_bold,
                              fontSize=16, textColor=colors.HexColor(BRAND["navy"]),
                              alignment=TA_CENTER, leading=20)
    kpi_delta = ParagraphStyle("kd", parent=styles["Normal"], fontName=_font_name,
                                fontSize=8, alignment=TA_CENTER, leading=10)
    footer = ParagraphStyle("f", parent=styles["Normal"], fontName=_font_name,
                             fontSize=7, textColor=colors.HexColor("#9CA3AF"),
                             alignment=TA_CENTER, leading=9)
    date_p = ParagraphStyle("dp", parent=styles["Normal"], fontName=_font_name,
                             fontSize=10, textColor=colors.HexColor(BRAND["grey_700"]),
                             leading=14)

    # Status color palette
    status_style = {
        "GOOD":  {"bg": "#ECFDF5", "fg": "#065F46", "border": "#A7F3D0"},
        "WATCH": {"bg": "#FFFBEB", "fg": "#92400E", "border": "#FDE68A"},
        "ALERT": {"bg": "#FEF2F2", "fg": "#991B1B", "border": "#FECACA"},
    }.get(overall_status.upper(), {"bg": BRAND["grey_50"], "fg": BRAND["grey_700"], "border": BRAND["grey_200"]})

    story = []
    story.extend(_branded_header("DIRECTOR BRIEF", _font_name, _font_bold, styles,
                                  mm, Table, TableStyle, Paragraph, Spacer,
                                  colors, ParagraphStyle))

    # ── Date + Overall Status Pill ──────────────────────────────────
    meta_style = ParagraphStyle("dm", parent=styles["Normal"], fontName=_font_name,
                                 fontSize=10, textColor=colors.HexColor(BRAND["grey_700"]),
                                 leading=14)
    status_pill_style = ParagraphStyle("sp", parent=styles["Normal"],
                                        fontName=_font_bold, fontSize=11,
                                        textColor=colors.HexColor(status_style["fg"]),
                                        alignment=TA_CENTER, leading=14)
    mood_style = ParagraphStyle("ms", parent=styles["Normal"], fontName=_font_name,
                                 fontSize=8.5, textColor=colors.HexColor(BRAND["grey_500"]),
                                 leading=11, spaceAfter=2)

    header_grid = Table([[
        Paragraph(f"<b>Date:</b> {date_str}<br/>"
                  f"<b>Prepared for:</b> {COMPANY['owner']}"
                  + (f"<br/><font color='{BRAND['grey_500']}' size='8'>{mood}</font>" if mood else ""),
                  meta_style),
        Table([[Paragraph(f"STATUS: {overall_status.upper()}", status_pill_style)]],
              colWidths=[55 * mm], rowHeights=[18 * mm]),
    ]], colWidths=[125 * mm, 55 * mm])
    header_grid.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
    ]))
    # Style the status pill
    inner_pill = header_grid._cellvalues[0][1]
    if hasattr(inner_pill, "setStyle"):
        inner_pill.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(status_style["bg"])),
            ("BOX",        (0, 0), (-1, -1), 1.2, colors.HexColor(status_style["border"])),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
    story.append(header_grid)
    story.append(Spacer(1, 8))

    # ── Top Action Band (amber) ─────────────────────────────────────
    if top_action:
        act_style = ParagraphStyle("act", parent=styles["Normal"],
                                    fontName=_font_bold, fontSize=10,
                                    textColor=colors.HexColor(BRAND["amber_fg"]),
                                    leading=14)
        act_sub_style = ParagraphStyle("acts", parent=styles["Normal"],
                                        fontName=_font_name, fontSize=7.5,
                                        textColor=colors.HexColor(BRAND["amber_fg"]),
                                        leading=10,
                                        textTransform=None)
        act_tbl = Table([[
            Paragraph("TOP PRIORITY TODAY", act_sub_style),
        ], [
            Paragraph(top_action, act_style),
        ]], colWidths=[180 * mm])
        act_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["amber_bg"])),
            ("BOX",        (0, 0), (-1, -1), 0.6, colors.HexColor("#FDE68A")),
            ("LINEBEFORE", (0, 0), (0, -1), 4,   colors.HexColor(BRAND["gold_dark"])),
            ("LEFTPADDING",(0, 0), (-1, -1), 14),
            ("RIGHTPADDING",(0, 0), (-1, -1), 14),
            ("TOPPADDING", (0, 0), (0, 0), 8),
            ("BOTTOMPADDING", (0, 0), (0, 0), 0),
            ("TOPPADDING", (0, 1), (0, 1), 0),
            ("BOTTOMPADDING", (0, 1), (0, 1), 8),
        ]))
        story.append(act_tbl)
        story.append(Spacer(1, 7))

    # ── KPI grid with colored borders ──────────────────────────────
    if kpis:
        # KPI pad style: delta color drives LEFT border color
        def _dir_indicator(direction, delta):
            if not direction:
                # Infer from delta prefix
                if delta.startswith("+") or "up" in delta.lower():
                    direction = "up"
                elif delta.startswith("-") or "down" in delta.lower() or "overdue" in delta.lower():
                    direction = "down"
                else:
                    direction = "flat"
            if direction == "up":
                return "\u25B2", "#16A34A"
            if direction == "down":
                return "\u25BC", "#DC2626"
            return "\u25CF", BRAND["grey_500"]

        per_row = 3
        # Build cards
        cards = []
        for k in kpis:
            label = k.get("label", "")
            value = k.get("value", "")
            delta = k.get("delta", "")
            direction = k.get("direction", "")
            arrow, arrow_color = _dir_indicator(direction, delta)
            border_color = k.get("color") or arrow_color

            card_label = Paragraph(label, kpi_label)
            card_val   = Paragraph(value, kpi_val)
            delta_html = (f"<font color='{arrow_color}'><b>{arrow}</b> {delta}</font>"
                          if delta else "")
            card_delta = Paragraph(delta_html, kpi_delta)

            card = Table([[card_label], [card_val], [card_delta]],
                          colWidths=[56 * mm])
            card.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["white"])),
                ("BOX",        (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
                ("LINEBEFORE", (0, 0), (0, -1), 4, colors.HexColor(border_color)),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            cards.append(card)
        while len(cards) % per_row != 0:
            cards.append("")
        for i in range(0, len(cards), per_row):
            row = cards[i:i + per_row]
            grid = Table([row], colWidths=[60 * mm] * per_row)
            grid.setStyle(TableStyle([
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            story.append(grid)
            story.append(Spacer(1, 4))
    story.append(Spacer(1, 4))

    # ── Narrative sections with colored accent bar ──────────────────
    # Use glyphs that Arial / DejaVu have reliably (tested: ▲ ▼ ● ■ •)
    default_accents = [
        ("\u25A0", BRAND["grey_500"]),   # ■ Yesterday (grey)
        ("\u25B2", "#4F46E5"),           # ▲ Today (indigo)
        ("\u2666", BRAND["gold"]),        # ♦ 15-day (gold diamond suit)
        ("\u25CF", "#DC2626"),            # ● Risks (red)
    ]
    for idx, sec in enumerate(sections):
        title = sec.get("title", "")
        accent = sec.get("accent") or default_accents[idx % len(default_accents)][1]
        icon   = sec.get("icon") or default_accents[idx % len(default_accents)][0]

        # Section header with accent bar
        h_tbl = Table([[
            Paragraph(f"<font color='{accent}'><b>{icon}</b></font> "
                      f"&nbsp;<font color='{BRAND['navy_light']}'><b>{title}</b></font>",
                      ParagraphStyle("sh2", parent=styles["Normal"],
                                      fontName=_font_bold, fontSize=11,
                                      leading=14))
        ]], colWidths=[180 * mm])
        h_tbl.setStyle(TableStyle([
            ("LINEBEFORE", (0, 0), (0, 0), 3, colors.HexColor(accent)),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(h_tbl)

        # Bullets (use • which every font has). Escape ampersands so
        # "L&T" doesn't turn into "L&T;" via accidental HTML entity.
        for b in sec.get("bullets", []):
            safe_b = b.replace("&", "&amp;")
            # Undo double-escape if caller already used &amp;
            safe_b = safe_b.replace("&amp;amp;", "&amp;")
            story.append(Paragraph(
                f"<font color='{accent}'><b>\u2022</b></font> &nbsp;{safe_b}",
                ParagraphStyle("bl2", parent=styles["Normal"],
                                fontName=_font_name, fontSize=9,
                                leading=12.5, leftIndent=14, spaceAfter=2)))
        story.append(Spacer(1, 2))

    # ── Signature / Read & Acknowledged band ────────────────────────
    story.append(Spacer(1, 4))
    ack_tbl = Table([[
        Paragraph(f"<b>Prepared by:</b> PPS AI Engine<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>Auto-generated {_today()} {_now_time()}</font>",
                  ParagraphStyle("ack1", parent=styles["Normal"],
                                  fontName=_font_name, fontSize=9, leading=12)),
        Paragraph(f"<b>Read &amp; Acknowledged</b><br/><br/>"
                  f"_______________________<br/>"
                  f"{COMPANY['owner']}<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>Director &middot; {COMPANY['short']}</font>",
                  ParagraphStyle("ack2", parent=styles["Normal"],
                                  fontName=_font_name, fontSize=9, leading=12)),
    ]], colWidths=[90 * mm, 90 * mm])
    ack_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(BRAND["grey_50"])),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(ack_tbl)

    # Footer
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"{COMPANY['name']} &middot; CIN {COMPANY['cin']} &middot; "
        f"Confidential — For {COMPANY['owner']} only",
        footer))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# Shared font registrar (used by the three new doc types below)
# ═══════════════════════════════════════════════════════════════════════════

def _register_unicode_font():
    """Register a Unicode TTF so ₹ glyph renders.
    Returns (_font_name, _font_bold, _rupee)."""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return "Helvetica", "Helvetica-Bold", "Rs."
    import os as _os
    for p, bp in [
        ("C:/Windows/Fonts/arial.ttf",                                 "C:/Windows/Fonts/arialbd.ttf"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/Library/Fonts/Arial.ttf",                                   "/Library/Fonts/Arial Bold.ttf"),
    ]:
        if _os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("MainFontX", p))
                if _os.path.exists(bp):
                    pdfmetrics.registerFont(TTFont("MainBoldX", bp))
                    return "MainFontX", "MainBoldX", "\u20b9"
                return "MainFontX", "MainFontX", "\u20b9"
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold", "Rs."


# ═══════════════════════════════════════════════════════════════════════════
# SALES ORDER
# ═══════════════════════════════════════════════════════════════════════════

def build_sales_order_pdf(customer_name: str, customer_address: str,
                          customer_gst: str,
                          so_no: str, so_date: str,
                          dispatch_date: str, dispatch_from: str,
                          items: list[dict]) -> bytes:
    """Premium Sales Order PDF (seller-side order confirmation)."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle)
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return b""

    _font_name, _font_bold, _rupee = _register_unicode_font()

    def _fmt_amt(n):
        return fmt_inr(n).lstrip("\u20b9")

    taxable = sum(it["qty"] * it["rate"] for it in items)
    gst = taxable * 0.18
    grand = taxable + gst

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15 * mm, rightMargin=15 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm,
                             title=f"Sales Order {so_no}",
                             author=COMPANY["name"])
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("son", parent=styles["Normal"], fontName=_font_name,
                             fontSize=9, leading=13)
    section_h = ParagraphStyle("sosh", parent=styles["Heading3"], fontName=_font_bold,
                                fontSize=10, textColor=colors.HexColor(BRAND["navy_light"]),
                                spaceAfter=4, spaceBefore=10, leading=14)
    footer = ParagraphStyle("sof", parent=styles["Normal"], fontName=_font_name,
                             fontSize=7, textColor=colors.HexColor("#9CA3AF"),
                             alignment=TA_CENTER, leading=9)

    story = []
    story.extend(_branded_header("SALES ORDER", _font_name, _font_bold, styles,
                                  mm, Table, TableStyle, Paragraph, Spacer,
                                  colors, ParagraphStyle))

    meta = Table([[
        Paragraph(f"<b>SO #:</b> {so_no}<br/>"
                  f"<b>SO Date:</b> {so_date}<br/>"
                  f"<b>Dispatch By:</b> {dispatch_date}<br/>"
                  f"<b>Dispatch From:</b> {dispatch_from}", normal),
        Paragraph(f"<b>Buyer:</b><br/><b>{customer_name}</b><br/>"
                  f"{customer_address}<br/>"
                  f"GSTIN: {customer_gst or 'Unregistered'}", normal),
    ]], colWidths=[85 * mm, 95 * mm])
    meta.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta)
    story.append(Spacer(1, 12))

    rows = [["#", "Description", "HSN", "Qty (MT)", f"Rate ({_rupee}/MT)", f"Amount ({_rupee})"]]
    for i, it in enumerate(items, 1):
        rows.append([str(i), it.get("desc", "Bitumen VG30"), COMPANY["hsn"],
                     f"{it['qty']:,.0f}", _fmt_amt(it["rate"]),
                     _fmt_amt(it["qty"] * it["rate"])])
    rows.append(["", "Sub-total", "", "", "", _fmt_amt(taxable)])
    rows.append(["", "GST @ 18%", "", "", "", _fmt_amt(gst)])
    rows.append(["", "Order Value", "", "", "", _fmt_amt(grand)])
    items_tbl = Table(rows,
                      colWidths=[10 * mm, 62 * mm, 22 * mm, 22 * mm, 28 * mm, 36 * mm])
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND["navy"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), _font_bold),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), _font_name),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(BRAND["amber_bg"])),
        ("FONTNAME", (0, -1), (-1, -1), _font_bold),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor(BRAND["amber_fg"])),
        ("FONTSIZE", (0, -1), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Order Terms", section_h))
    terms = [
        "Rates are ex-terminal / ex-warehouse unless stated otherwise.",
        "Payment: 100% advance against pro-forma invoice before dispatch.",
        f"Dispatch scheduled on / before {dispatch_date} subject to payment.",
        "Goods conform to IS 73:2018; test certificate per tanker.",
        "Transit risk buyer's account post dispatch; insurance available on request.",
        "All disputes subject to Vadodara, Gujarat jurisdiction.",
    ]
    for i, t in enumerate(terms, 1):
        story.append(Paragraph(
            f"<font color='{BRAND['navy_light']}'><b>{i}.</b></font> &nbsp;{t}",
            ParagraphStyle("sot", parent=styles["Normal"], fontName=_font_name,
                            fontSize=9, leading=13, leftIndent=14, spaceAfter=3)))
    story.append(Spacer(1, 8))

    # Bank block (payment target)
    bank = Table([[Paragraph(
        f"<b>Remit advance to:</b> {COMPANY['bank']} &middot; "
        f"A/C <b>{COMPANY['bank_ac']}</b> &middot; IFSC <b>{COMPANY['bank_ifsc']}</b> &middot; "
        f"Beneficiary: {COMPANY['name']}", normal)]],
        colWidths=[180 * mm])
    bank.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["grey_50"])),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(bank)
    story.append(Spacer(1, 10))

    # Signature
    sig = Table([[
        Paragraph(f"<b>For {COMPANY['name']} (Seller)</b><br/><br/><br/>"
                  f"_______________________<br/>"
                  f"Authorized Signatory<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>{COMPANY['owner']} &middot; {COMPANY['phone']}</font>",
                  normal),
        Paragraph(f"<b>Accepted by Buyer</b><br/><br/><br/>"
                  f"_______________________<br/>"
                  f"Signature &middot; Stamp<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>{customer_name}</font>",
                  normal),
    ]], colWidths=[90 * mm, 90 * mm])
    sig.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(sig)
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"{COMPANY['name']} &middot; GST {COMPANY['gst']} &middot; "
        f"{COMPANY['city']} &middot; SO confidential — For {customer_name} only",
        footer))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# PAYMENT RECEIPT
# ═══════════════════════════════════════════════════════════════════════════

def build_payment_receipt_pdf(party_name: str, party_address: str,
                              receipt_no: str, receipt_date: str,
                              amount: float, mode: str,
                              reference_no: str = "",
                              against_invoice: str = "",
                              direction: str = "received") -> bytes:
    """
    Premium Payment Receipt PDF.
    direction = "received" (money in from customer) or "paid" (money out to supplier).
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle)
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return b""

    _font_name, _font_bold, _rupee = _register_unicode_font()
    is_in = direction == "received"
    title_label = "PAYMENT RECEIPT" if is_in else "PAYMENT ADVICE"
    flow_label  = "Received From" if is_in else "Paid To"
    amount_tag  = "Amount Received" if is_in else "Amount Paid"

    def _fmt_amt(n):
        return fmt_inr(n).lstrip("\u20b9")

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15 * mm, rightMargin=15 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm,
                             title=f"{title_label} {receipt_no}",
                             author=COMPANY["name"])
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("pn", parent=styles["Normal"], fontName=_font_name,
                             fontSize=9.5, leading=14)
    big = ParagraphStyle("pbig", parent=styles["Normal"], fontName=_font_bold,
                          fontSize=22, textColor=colors.HexColor(BRAND["navy"]),
                          alignment=TA_CENTER, leading=26)
    tag = ParagraphStyle("ptag", parent=styles["Normal"], fontName=_font_name,
                          fontSize=9, textColor=colors.HexColor(BRAND["grey_500"]),
                          alignment=TA_CENTER, leading=12)
    footer = ParagraphStyle("pf", parent=styles["Normal"], fontName=_font_name,
                             fontSize=7, textColor=colors.HexColor("#9CA3AF"),
                             alignment=TA_CENTER, leading=9)

    story = []
    story.extend(_branded_header(title_label, _font_name, _font_bold, styles,
                                  mm, Table, TableStyle, Paragraph, Spacer,
                                  colors, ParagraphStyle))

    # Meta
    meta = Table([[
        Paragraph(f"<b>Receipt #:</b> {receipt_no}<br/>"
                  f"<b>Date:</b> {receipt_date}<br/>"
                  f"<b>Mode:</b> {mode}"
                  + (f"<br/><b>Ref #:</b> {reference_no}" if reference_no else ""),
                  normal),
        Paragraph(f"<b>{flow_label}:</b><br/><b>{party_name}</b><br/>"
                  f"{party_address}"
                  + (f"<br/><b>Against Invoice:</b> {against_invoice}" if against_invoice else ""),
                  normal),
    ]], colWidths=[85 * mm, 95 * mm])
    meta.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(meta)
    story.append(Spacer(1, 14))

    # Big amount band
    amt_tbl = Table([
        [Paragraph(amount_tag, tag)],
        [Paragraph(f"{_rupee}{_fmt_amt(amount)}", big)],
        [Paragraph(_amount_in_words(amount) + " Only", tag)],
    ], colWidths=[180 * mm])
    amt_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["amber_bg"])),
        ("BOX",        (0, 0), (-1, -1), 0.8, colors.HexColor(BRAND["gold"])),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(amt_tbl)
    story.append(Spacer(1, 16))

    # Thank-you / advice band
    msg = ("Thank you for your prompt payment. Your account stands updated."
           if is_in else
           "Payment released per agreed terms. Kindly acknowledge receipt.")
    story.append(Paragraph(
        f"<i>{msg}</i>",
        ParagraphStyle("pmsg", parent=styles["Normal"], fontName=_font_name,
                        fontSize=10, textColor=colors.HexColor(BRAND["navy_light"]),
                        alignment=TA_CENTER, leading=14)))
    story.append(Spacer(1, 22))

    # Signature
    sig = Table([[
        Paragraph(f"<b>For {COMPANY['name']}</b><br/><br/><br/>"
                  f"_______________________<br/>"
                  f"Authorized Signatory<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>{COMPANY['owner']} &middot; {COMPANY['phone']}</font>",
                  normal),
    ]], colWidths=[180 * mm])
    sig.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 110),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(sig)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"{COMPANY['name']} &middot; GST {COMPANY['gst']} &middot; "
        f"Bank: {COMPANY['bank']} A/C {COMPANY['bank_ac']} &middot; {COMPANY['city']}",
        footer))

    doc.build(story)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# DELIVERY NOTE / DISPATCH CHALLAN
# ═══════════════════════════════════════════════════════════════════════════

def build_delivery_note_pdf(customer_name: str, delivery_address: str,
                            dn_no: str, dn_date: str,
                            dispatch_from: str,
                            vehicle_no: str, driver_name: str,
                            driver_phone: str,
                            items: list[dict],
                            so_ref: str = "",
                            invoice_ref: str = "") -> bytes:
    """Premium Delivery Note / Dispatch Challan PDF."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                         Table, TableStyle)
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        return b""

    _font_name, _font_bold, _rupee = _register_unicode_font()

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=15 * mm, rightMargin=15 * mm,
                             topMargin=12 * mm, bottomMargin=12 * mm,
                             title=f"Delivery Note {dn_no}",
                             author=COMPANY["name"])
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("dnn", parent=styles["Normal"], fontName=_font_name,
                             fontSize=9, leading=13)
    section_h = ParagraphStyle("dnsh", parent=styles["Heading3"], fontName=_font_bold,
                                fontSize=10, textColor=colors.HexColor(BRAND["navy_light"]),
                                spaceAfter=4, spaceBefore=10, leading=14)
    footer = ParagraphStyle("dnf", parent=styles["Normal"], fontName=_font_name,
                             fontSize=7, textColor=colors.HexColor("#9CA3AF"),
                             alignment=TA_CENTER, leading=9)

    story = []
    story.extend(_branded_header("DELIVERY NOTE", _font_name, _font_bold, styles,
                                  mm, Table, TableStyle, Paragraph, Spacer,
                                  colors, ParagraphStyle))

    meta = Table([[
        Paragraph(f"<b>DN #:</b> {dn_no}<br/>"
                  f"<b>Date:</b> {dn_date}<br/>"
                  f"<b>Dispatched From:</b> {dispatch_from}"
                  + (f"<br/><b>SO Ref:</b> {so_ref}" if so_ref else "")
                  + (f"<br/><b>Invoice:</b> {invoice_ref}" if invoice_ref else ""),
                  normal),
        Paragraph(f"<b>Consignee:</b><br/><b>{customer_name}</b><br/>"
                  f"{delivery_address}", normal),
    ]], colWidths=[85 * mm, 95 * mm])
    meta.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.5, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(meta)
    story.append(Spacer(1, 10))

    # Vehicle block
    veh = Table([[
        Paragraph(f"<b>Vehicle:</b> {vehicle_no}", normal),
        Paragraph(f"<b>Driver:</b> {driver_name}", normal),
        Paragraph(f"<b>Driver Phone:</b> {driver_phone}", normal),
    ]], colWidths=[60 * mm, 60 * mm, 60 * mm])
    veh.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(BRAND["grey_50"])),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(veh)
    story.append(Spacer(1, 12))

    # Items
    rows = [["#", "Description", "HSN", "Qty (MT)", "Batch / Remark"]]
    total_qty = 0.0
    for i, it in enumerate(items, 1):
        q = float(it.get("qty", 0))
        total_qty += q
        rows.append([str(i), it.get("desc", "Bitumen VG30"), COMPANY["hsn"],
                     f"{q:,.2f}", it.get("batch", "")])
    rows.append(["", "Total Qty", "", f"{total_qty:,.2f}", ""])
    items_tbl = Table(rows,
                      colWidths=[10 * mm, 72 * mm, 22 * mm, 26 * mm, 50 * mm])
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(BRAND["navy"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), _font_bold),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 1), (-1, -1), _font_name),
        ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(BRAND["amber_bg"])),
        ("FONTNAME", (0, -1), (-1, -1), _font_bold),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor(BRAND["amber_fg"])),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(items_tbl)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Dispatch Terms", section_h))
    terms = [
        "Goods dispatched in good condition per IS 73:2018. Quality certificate attached.",
        "Consignee to verify quantity & condition at the time of unloading.",
        "Any shortage / damage to be reported within 24 hours of receipt.",
        "Transit risk as per agreed SO terms; insurance covered only if stated.",
        "This delivery note is for dispatch reference and not a tax invoice.",
    ]
    for i, t in enumerate(terms, 1):
        story.append(Paragraph(
            f"<font color='{BRAND['navy_light']}'><b>{i}.</b></font> &nbsp;{t}",
            ParagraphStyle("dnt", parent=styles["Normal"], fontName=_font_name,
                            fontSize=9, leading=13, leftIndent=14, spaceAfter=3)))
    story.append(Spacer(1, 10))

    # Receipt block
    rec = Table([[
        Paragraph(f"<b>For {COMPANY['name']} (Dispatcher)</b><br/><br/><br/>"
                  f"_______________________<br/>"
                  f"Authorized Signatory", normal),
        Paragraph(f"<b>Received by Consignee</b><br/><br/><br/>"
                  f"_______________________<br/>"
                  f"Signature &middot; Stamp &middot; Date<br/>"
                  f"<font color='{BRAND['grey_500']}' size='8'>{customer_name}</font>",
                  normal),
    ]], colWidths=[90 * mm, 90 * mm])
    rec.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.4, colors.HexColor(BRAND["grey_200"])),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(rec)
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"{COMPANY['name']} &middot; GST {COMPANY['gst']} &middot; "
        f"{COMPANY['city']} &middot; Document confidential",
        footer))

    doc.build(story)
    return buf.getvalue()
