"""
universal_action_engine.py — Universal Send / Export Engine for PPS Anantam
===========================================================================
Transforms any page's content (KPIs, tables, charts, insights) into
Email / WhatsApp / PDF / CSV / Excel format. Zero-cost — uses only
Python stdlib + already-installed packages (openpyxl, reportlab).

Usage:
    from universal_action_engine import PageContext, build_email_report, build_excel_export
    ctx = PageContext("Price Prediction", "VG30 rising 3%", ...)
    report = build_email_report(ctx)
"""

from __future__ import annotations

import csv
import io
import json
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta

IST = timezone(timedelta(hours=5, minutes=30))
COMPANY = "PPS Anantam Agentic AI Eco System"

# ── Vastu Design Colors ──────────────────────────────────────────────────────
NAVY = "#1e3a5f"
IVORY = "#faf7f2"
GOLD = "#c9a84c"
GREEN = "#2d6a4f"
FIRE = "#b85c38"
CREAM = "#f0ebe1"


def _now_ist() -> str:
    return datetime.now(IST).strftime("%Y-%m-%d %H:%M IST")


def _now_ist_filename() -> str:
    return datetime.now(IST).strftime("%Y-%m-%d_%H%M")


# ═══════════════════════════════════════════════════════════════════════════
# PAGE CONTEXT — universal data container for any page
# ═══════════════════════════════════════════════════════════════════════════

@dataclass
class PageContext:
    page_name: str = ""
    summary_text: str = ""
    kpis: list = field(default_factory=list)           # [(label, value, delta), ...]
    tables: list = field(default_factory=list)          # [{"title": str, "headers": [...], "rows": [[...]]}]
    chart_figures: list = field(default_factory=list)   # [{"fig": plotly_fig, "caption": str}]
    filters: dict = field(default_factory=dict)         # {"Grade": "VG30", ...}
    insights: list = field(default_factory=list)        # ["insight1", ...]
    action_items: list = field(default_factory=list)    # ["action1", ...]
    timestamp_ist: str = field(default_factory=_now_ist)


# ═══════════════════════════════════════════════════════════════════════════
# EMAIL REPORT BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def build_email_report(ctx: PageContext) -> dict:
    """
    Build HTML email from page context.
    Returns {"subject": str, "body_html": str, "body_text": str}
    """
    ts = ctx.timestamp_ist or _now_ist()
    subject = f"[PPS Anantam] {ctx.page_name} Report — {ts}"

    # Build HTML
    html_parts = []

    # Header banner
    html_parts.append(f"""
    <div style="background:{NAVY};color:#ffffff;padding:16px 24px;border-radius:8px 8px 0 0;
                font-family:'Segoe UI',Arial,sans-serif;">
      <div style="font-size:18px;font-weight:700;">{COMPANY}</div>
      <div style="font-size:14px;color:{GOLD};margin-top:4px;">{ctx.page_name} Report</div>
      <div style="font-size:11px;color:#94a3b8;margin-top:2px;">{ts}</div>
    </div>
    """)

    # Filters
    if ctx.filters:
        filter_str = " | ".join(f"<b>{k}:</b> {v}" for k, v in ctx.filters.items())
        html_parts.append(f"""
        <div style="background:{CREAM};padding:8px 16px;font-size:12px;color:#475569;
                    border-bottom:1px solid #e2e8f0;">
          Filters: {filter_str}
        </div>
        """)

    # Body container
    html_parts.append(f'<div style="padding:16px 24px;font-family:\'Segoe UI\',Arial,sans-serif;background:#ffffff;color:#1e293b;">')

    # Summary
    if ctx.summary_text:
        html_parts.append(f'<p style="font-size:14px;line-height:1.6;margin-bottom:16px;">{ctx.summary_text}</p>')

    # KPIs
    if ctx.kpis:
        kpi_cells = []
        for item in ctx.kpis:
            label = item[0] if len(item) > 0 else ""
            value = item[1] if len(item) > 1 else ""
            delta = item[2] if len(item) > 2 else ""
            delta_color = GREEN if delta and not str(delta).startswith("-") else FIRE if delta else "#94a3b8"
            kpi_cells.append(f"""
            <td style="text-align:center;padding:12px 16px;vertical-align:top;">
              <div style="font-size:11px;color:#64748b;text-transform:uppercase;">{label}</div>
              <div style="font-size:20px;font-weight:700;color:{NAVY};">{value}</div>
              <div style="font-size:11px;color:{delta_color};">{delta}</div>
            </td>
            """)
        html_parts.append(f"""
        <table style="width:100%;border-collapse:collapse;margin-bottom:16px;background:{IVORY};border-radius:6px;">
          <tr>{''.join(kpi_cells)}</tr>
        </table>
        """)

    # Tables
    for tbl in ctx.tables:
        title = tbl.get("title", "Data")
        headers = tbl.get("headers", [])
        rows = tbl.get("rows", [])
        html_parts.append(f'<h3 style="font-size:14px;color:{NAVY};margin:16px 0 8px 0;">{title}</h3>')
        if headers and rows:
            th_cells = "".join(
                f'<th style="background:{NAVY};color:#fff;padding:6px 10px;font-size:11px;text-align:left;'
                f'border:1px solid #334155;">{h}</th>'
                for h in headers
            )
            tr_rows = []
            for i, row in enumerate(rows[:50]):  # limit to 50 rows in email
                bg = "#ffffff" if i % 2 == 0 else IVORY
                td_cells = "".join(
                    f'<td style="padding:5px 10px;font-size:11px;border:1px solid #e2e8f0;background:{bg};">{c}</td>'
                    for c in row
                )
                tr_rows.append(f"<tr>{td_cells}</tr>")
            html_parts.append(f"""
            <table style="width:100%;border-collapse:collapse;margin-bottom:12px;">
              <thead><tr>{th_cells}</tr></thead>
              <tbody>{''.join(tr_rows)}</tbody>
            </table>
            """)
            if len(rows) > 50:
                html_parts.append(f'<p style="font-size:11px;color:#94a3b8;">Showing 50 of {len(rows)} rows. View full data in dashboard.</p>')

    # Insights
    if ctx.insights:
        html_parts.append(f'<h3 style="font-size:14px;color:{NAVY};margin:16px 0 8px 0;">Key Insights</h3><ul style="margin:0;padding-left:20px;">')
        for ins in ctx.insights:
            html_parts.append(f'<li style="font-size:12px;margin-bottom:4px;">{ins}</li>')
        html_parts.append("</ul>")

    # Action Items
    if ctx.action_items:
        html_parts.append(f'<h3 style="font-size:14px;color:{NAVY};margin:16px 0 8px 0;">Action Items</h3><ol style="margin:0;padding-left:20px;">')
        for act in ctx.action_items:
            html_parts.append(f'<li style="font-size:12px;font-weight:600;margin-bottom:4px;">{act}</li>')
        html_parts.append("</ol>")

    # Close body container
    html_parts.append("</div>")

    # Footer
    html_parts.append(f"""
    <div style="background:#f1f5f9;padding:12px 24px;border-radius:0 0 8px 8px;
                font-size:10px;color:#64748b;border-top:1px solid #e2e8f0;">
      <div>CONFIDENTIAL — {COMPANY} | Vadodara, Gujarat</div>
      <div>Generated: {ts} | Source: PPS Anantam Dashboard</div>
      <div style="margin-top:4px;">This is an automated report. Do not reply to this email.</div>
    </div>
    """)

    body_html = "\n".join(html_parts)

    # Plain text fallback
    text_lines = [
        f"{COMPANY}",
        f"{ctx.page_name} Report — {ts}",
        "",
    ]
    if ctx.summary_text:
        text_lines.extend([ctx.summary_text, ""])
    if ctx.kpis:
        for item in ctx.kpis:
            label = item[0] if len(item) > 0 else ""
            value = item[1] if len(item) > 1 else ""
            text_lines.append(f"  {label}: {value}")
        text_lines.append("")
    if ctx.insights:
        text_lines.append("KEY INSIGHTS:")
        for ins in ctx.insights:
            text_lines.append(f"  - {ins}")
        text_lines.append("")
    if ctx.action_items:
        text_lines.append("ACTION ITEMS:")
        for i, act in enumerate(ctx.action_items, 1):
            text_lines.append(f"  {i}. {act}")
    body_text = "\n".join(text_lines)

    return {"subject": subject, "body_html": body_html, "body_text": body_text}


# ═══════════════════════════════════════════════════════════════════════════
# WHATSAPP SUMMARY BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def build_whatsapp_summary(ctx: PageContext) -> str:
    """Build short text summary for WhatsApp (max 4096 chars)."""
    ts = ctx.timestamp_ist or _now_ist()
    lines = [
        f"*{COMPANY}*",
        f"*{ctx.page_name}*",
        f"_{ts}_",
        "",
    ]

    if ctx.summary_text:
        lines.extend([ctx.summary_text, ""])

    # KPIs
    if ctx.kpis:
        for item in ctx.kpis:
            label = item[0] if len(item) > 0 else ""
            value = item[1] if len(item) > 1 else ""
            delta = item[2] if len(item) > 2 else ""
            delta_str = f" ({delta})" if delta else ""
            lines.append(f"{label}: *{value}*{delta_str}")
        lines.append("")

    # Insights (top 5)
    if ctx.insights:
        lines.append("*Key Insights:*")
        for ins in ctx.insights[:5]:
            lines.append(f"- {ins}")
        lines.append("")

    # Action items (top 3)
    if ctx.action_items:
        lines.append("*Action Items:*")
        for i, act in enumerate(ctx.action_items[:3], 1):
            lines.append(f"{i}. {act}")
        lines.append("")

    lines.append("_Sent via PPS Dashboard_")

    result = "\n".join(lines)
    if len(result) > 4096:
        result = result[:4090] + "\n..."
    return result


# ═══════════════════════════════════════════════════════════════════════════
# CSV EXPORT BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def build_csv_export(ctx: PageContext) -> bytes:
    """Build CSV from all tables in context. UTF-8-sig for Excel compatibility."""
    try:
        encoding = "utf-8-sig"
        from settings_engine import get as get_setting
        encoding = get_setting("csv_export_encoding", "utf-8-sig")
    except Exception:
        encoding = "utf-8-sig"

    output = io.StringIO()
    writer = csv.writer(output)

    # Metadata header
    writer.writerow(["Report", ctx.page_name])
    writer.writerow(["Generated", ctx.timestamp_ist or _now_ist()])
    if ctx.filters:
        writer.writerow(["Filters", json.dumps(ctx.filters, ensure_ascii=False)])
    writer.writerow([])

    # KPIs
    if ctx.kpis:
        writer.writerow(["--- KPIs ---"])
        writer.writerow(["Metric", "Value", "Change"])
        for item in ctx.kpis:
            label = item[0] if len(item) > 0 else ""
            value = item[1] if len(item) > 1 else ""
            delta = item[2] if len(item) > 2 else ""
            writer.writerow([label, value, delta])
        writer.writerow([])

    # Tables
    for tbl in ctx.tables:
        title = tbl.get("title", "Data")
        headers = tbl.get("headers", [])
        rows = tbl.get("rows", [])
        writer.writerow([f"--- {title} ---"])
        if headers:
            writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        writer.writerow([])

    return output.getvalue().encode(encoding)


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def build_excel_export(ctx: PageContext) -> bytes:
    """Build Excel workbook with metadata sheet + data sheets. Uses openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Vastu styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="1E293B")
    kpi_fill = PatternFill(start_color="FAF7F2", end_color="FAF7F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # ── Sheet 1: Report Metadata ──────────────────────────────────────────
    ws_meta = wb.active
    ws_meta.title = "Report Metadata"

    meta_rows = [
        ("Report", ctx.page_name),
        ("Generated", ctx.timestamp_ist or _now_ist()),
        ("Company", COMPANY),
    ]
    if ctx.filters:
        for k, v in ctx.filters.items():
            meta_rows.append((f"Filter: {k}", str(v)))
    if ctx.summary_text:
        meta_rows.append(("Summary", ctx.summary_text))

    for r_idx, (label, value) in enumerate(meta_rows, 1):
        cell_a = ws_meta.cell(row=r_idx, column=1, value=label)
        cell_a.font = Font(name="Segoe UI", size=10, bold=True, color="1E3A5F")
        cell_b = ws_meta.cell(row=r_idx, column=2, value=value)
        cell_b.font = data_font

    # KPIs
    if ctx.kpis:
        kpi_start = len(meta_rows) + 2
        ws_meta.cell(row=kpi_start, column=1, value="KPIs").font = Font(name="Segoe UI", size=11, bold=True, color="1E3A5F")
        kpi_headers = ["Metric", "Value", "Change"]
        for c_idx, h in enumerate(kpi_headers, 1):
            cell = ws_meta.cell(row=kpi_start + 1, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for i, item in enumerate(ctx.kpis):
            row_num = kpi_start + 2 + i
            for c_idx in range(3):
                val = item[c_idx] if c_idx < len(item) else ""
                cell = ws_meta.cell(row=row_num, column=c_idx + 1, value=val)
                cell.font = data_font
                cell.fill = kpi_fill
                cell.border = thin_border

    # Auto-width for metadata sheet
    for col_cells in ws_meta.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws_meta.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # ── Data Sheets (one per table) ───────────────────────────────────────
    for tbl_idx, tbl in enumerate(ctx.tables):
        title = tbl.get("title", f"Data {tbl_idx + 1}")
        headers = tbl.get("headers", [])
        rows = tbl.get("rows", [])

        # Sanitize sheet name (max 31 chars, no special chars)
        safe_name = "".join(c for c in title if c.isalnum() or c in " _-")[:31] or f"Sheet{tbl_idx + 2}"
        ws = wb.create_sheet(title=safe_name)

        # Headers
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = data_font
                cell.border = thin_border

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-width
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# PDF REPORT BUILDER (delegates to existing engine)
# ═══════════════════════════════════════════════════════════════════════════

def build_pdf_report(ctx: PageContext, orientation: str = "portrait") -> bytes:
    """Build PDF using existing pdf_export_engine. Converts PageContext to sections."""
    from pdf_export_engine import build_generic_pdf

    sections = []

    # Summary
    if ctx.summary_text:
        sections.append({"type": "paragraph", "text": ctx.summary_text})

    # KPIs
    if ctx.kpis:
        sections.append({"type": "kpis", "items": ctx.kpis})

    # Tables
    for tbl in ctx.tables:
        sections.append({"type": "section", "text": tbl.get("title", "Data")})
        sections.append({
            "type": "table",
            "headers": tbl.get("headers", []),
            "rows": tbl.get("rows", []),
        })

    # Insights
    if ctx.insights:
        sections.append({"type": "section", "text": "Key Insights"})
        for ins in ctx.insights:
            sections.append({"type": "paragraph", "text": f"  - {ins}"})

    # Action items
    if ctx.action_items:
        sections.append({"type": "section", "text": "Action Items"})
        for i, act in enumerate(ctx.action_items, 1):
            sections.append({"type": "paragraph", "text": f"  {i}. {act}"})

    # Fallback if empty
    if not sections:
        from pdf_export_engine import _ts_ist
        sections = [
            {"type": "section", "text": "Page Overview"},
            {"type": "paragraph", "text": f"Report exported from {ctx.page_name} page."},
            {"type": "paragraph", "text": f"Generated: {_ts_ist()}"},
        ]

    return build_generic_pdf(
        page_title=ctx.page_name,
        sections=sections,
        filters=ctx.filters if ctx.filters else None,
        role="Admin",
        orientation=orientation,
    )


# ═══════════════════════════════════════════════════════════════════════════
# SEND VIA EMAIL
# ═══════════════════════════════════════════════════════════════════════════

def send_email_report(ctx: PageContext, recipients: list,
                      cc: str = "", subject_prefix: str = "") -> dict:
    """
    Send email report to recipients list.
    Returns {"queued": int, "failed": int, "queue_ids": list}
    """
    report = build_email_report(ctx)
    subject = f"{subject_prefix}{report['subject']}" if subject_prefix else report["subject"]

    result = {"queued": 0, "failed": 0, "queue_ids": []}

    try:
        from email_engine import EmailEngine
        ee = EmailEngine()

        for recip in recipients:
            try:
                qid = ee.queue_email(
                    to_email=recip.strip(),
                    subject=subject,
                    body_html=report["body_html"],
                    email_type="report",
                    cc=cc or None,
                )
                result["queued"] += 1
                result["queue_ids"].append(qid)
            except Exception:
                result["failed"] += 1
    except ImportError:
        result["failed"] = len(recipients)

    # Audit log
    try:
        from database import insert_audit_log
        from role_engine import get_current_username
        insert_audit_log({
            "username": get_current_username(),
            "action": "send_email_report",
            "resource": ctx.page_name,
            "details": json.dumps({"recipients": len(recipients), "queued": result["queued"]}),
        })
    except Exception:
        pass

    return result


# ═══════════════════════════════════════════════════════════════════════════
# SEND VIA WHATSAPP
# ═══════════════════════════════════════════════════════════════════════════

def send_whatsapp_report(ctx: PageContext, recipients: list) -> dict:
    """Send WhatsApp summary to phone numbers. Returns {"queued", "failed"}."""
    summary = build_whatsapp_summary(ctx)
    result = {"queued": 0, "failed": 0}

    try:
        from whatsapp_engine import WhatsAppEngine
        we = WhatsAppEngine()

        for phone in recipients:
            try:
                we.queue_message(
                    to_number=phone.strip(),
                    message_type="session",
                    session_text=summary,
                )
                result["queued"] += 1
            except Exception:
                result["failed"] += 1
    except ImportError:
        result["failed"] = len(recipients)

    # Audit log
    try:
        from database import insert_audit_log
        from role_engine import get_current_username
        insert_audit_log({
            "username": get_current_username(),
            "action": "send_whatsapp_report",
            "resource": ctx.page_name,
            "details": json.dumps({"recipients": len(recipients), "queued": result["queued"]}),
        })
    except Exception:
        pass

    return result


# ═══════════════════════════════════════════════════════════════════════════
# RECIPIENT RESOLUTION
# ═══════════════════════════════════════════════════════════════════════════

def resolve_recipients(list_name: str, channel: str = "email") -> list:
    """
    Resolve named recipient list to actual addresses.
    channel: 'email' returns email addresses, 'whatsapp' returns phone numbers
    """
    try:
        from database import get_recipient_lists
        all_lists = get_recipient_lists()
        for rl in all_lists:
            if rl.get("list_name", "").lower() == list_name.lower():
                try:
                    recipients = json.loads(rl.get("recipients", "[]"))
                except (json.JSONDecodeError, TypeError):
                    recipients = []
                return [
                    r.get(channel, r.get("email", ""))
                    for r in recipients
                    if r.get(channel) or r.get("email")
                ]
    except Exception:
        pass
    return []


# ═══════════════════════════════════════════════════════════════════════════
# SHARE LINK GENERATOR
# ═══════════════════════════════════════════════════════════════════════════

def generate_share_link(ctx: PageContext) -> str:
    """Generate internal share link (session-state based UUID key)."""
    try:
        import streamlit as st
        share_id = str(uuid.uuid4())[:8]
        key = f"_share_{share_id}"
        st.session_state[key] = {
            "page_name": ctx.page_name,
            "filters": ctx.filters,
            "timestamp": ctx.timestamp_ist,
        }
        return f"?share={share_id}&page={ctx.page_name.replace(' ', '+')}"
    except Exception:
        return f"?page={ctx.page_name.replace(' ', '+')}"


# ═══════════════════════════════════════════════════════════════════════════
# GENERIC CONTEXT BUILDER (for pages without custom context)
# ═══════════════════════════════════════════════════════════════════════════

def build_generic_context(page_name: str) -> PageContext:
    """Build a minimal PageContext for pages that don't provide their own."""
    return PageContext(
        page_name=page_name,
        summary_text=f"Report exported from the {page_name} page of {COMPANY}.",
        kpis=[],
        tables=[],
        chart_figures=[],
        filters={},
        insights=["View the live dashboard for interactive charts and full data."],
        action_items=[],
    )
